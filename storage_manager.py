import logging
import shutil
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@")


class ExcelStorageManager:
    def __init__(self, output_path: str, sheet_name: str = "Daily_Trend", verbose: bool = True):
        self.output_path = Path(output_path)
        self.sheet_name = sheet_name
        self.verbose = verbose
        self.columns = [
            "게시일시",
            "일자",
            "이름",
            "주요내용",
            "출처",
            "URL",
            "구분",
            "기술분류",
            "원문(옵션)",
        ]
        self.logger = logging.getLogger(self.__class__.__name__)

    def merge_and_save(self, new_rows: list[dict[str, str]]) -> tuple[int, int]:
        existing_df = self._read_existing()

        if not new_rows:
            if self.verbose:
                self.logger.info("no new rows to save")
            if not self.output_path.exists():
                self._save_to_excel(pd.DataFrame(columns=self.columns))
            return 0, len(existing_df)

        new_df = pd.DataFrame(new_rows)
        for col in self.columns:
            if col not in new_df.columns:
                new_df[col] = ""
        new_df = new_df[self.columns]

        existing_urls = self._normalized_url_set(existing_df["URL"]) if "URL" in existing_df.columns else set()
        new_df["_url_norm"] = new_df["URL"].map(self._normalize_url)
        if existing_urls:
            new_df = new_df[~new_df["_url_norm"].isin(existing_urls)]
        new_df = new_df.drop_duplicates(subset=["_url_norm"], keep="first")
        new_df = new_df.drop(columns=["_url_norm"])

        if new_df.empty:
            if self.verbose:
                self.logger.info("all rows are duplicates")
            return 0, len(existing_df)

        merged_df = pd.concat([existing_df, new_df], ignore_index=True)
        merged_df = self._sort_rows_desc(merged_df)
        merged_df = self._sanitize_for_excel(merged_df)
        self._save_to_excel(merged_df)

        added_count = len(new_df)
        total_count = len(merged_df)
        if self.verbose:
            self.logger.info(f"saved {added_count} new rows (total {total_count})")
        return added_count, total_count

    def _read_existing(self) -> pd.DataFrame:
        if not self.output_path.exists():
            return pd.DataFrame(columns=self.columns)

        try:
            df = pd.read_excel(self.output_path, sheet_name=self.sheet_name)
            for col in self.columns:
                if col not in df.columns:
                    df[col] = ""
            return df[self.columns]
        except Exception as exc:
            if self.verbose:
                self.logger.warning(f"failed to read existing file: {exc.__class__.__name__}: {exc}")
            return pd.DataFrame(columns=self.columns)

    def _save_to_excel(self, df: pd.DataFrame) -> None:
        if self.output_path.exists():
            self._rotate_backups()

        tmp_path = self.output_path.with_name(f"{self.output_path.stem}.tmp{self.output_path.suffix}")
        for attempt in range(3):
            try:
                with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="w") as writer:
                    df.to_excel(writer, sheet_name=self.sheet_name, index=False)

                self._style_excel(tmp_path)
                tmp_path.replace(self.output_path)
                return
            except PermissionError:
                if attempt < 2:
                    self.logger.warning(f"file permission error, retrying in 2s... ({attempt + 1}/3)")
                    time.sleep(2)
                else:
                    self.logger.error("failed to save excel due to permission error (is the file open?)")
                    raise
            except Exception as exc:
                self.logger.error(f"failed to save excel: {exc.__class__.__name__}: {exc}")
                raise
            finally:
                if tmp_path.exists():
                    tmp_path.unlink(missing_ok=True)

    def _rotate_backups(self) -> None:
        bak1 = self.output_path.with_suffix(".xlsx.bak.1")
        bak2 = self.output_path.with_suffix(".xlsx.bak.2")
        bak3 = self.output_path.with_suffix(".xlsx.bak.3")

        try:
            if bak2.exists():
                shutil.move(str(bak2), str(bak3))
            if bak1.exists():
                shutil.move(str(bak1), str(bak2))
            shutil.copy2(self.output_path, bak1)
        except Exception as exc:
            if self.verbose:
                self.logger.warning(f"backup rotation failed: {exc.__class__.__name__}: {exc}")

    def _style_excel(self, path: Path) -> None:
        try:
            wb = load_workbook(path)
            ws = wb[self.sheet_name]

            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            header_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            column_widths = {
                "A": 20,
                "B": 12,
                "C": 15,
                "D": 80,
                "E": 10,
                "F": 10,
                "G": 10,
                "H": 10,
                "I": 10,
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            max_row = ws.max_row
            if max_row > 1:
                for row_idx in range(2, max_row + 1):
                    ws.cell(row=row_idx, column=4).alignment = left_align
                    for col_idx in [1, 2, 3, 5, 7, 8]:
                        ws.cell(row=row_idx, column=col_idx).alignment = center_align

            wb.save(path)
        except Exception as exc:
            if self.verbose:
                self.logger.warning(f"excel styling failed: {exc.__class__.__name__}: {exc}")

    def _sort_rows_desc(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        if dataframe.empty:
            return dataframe[self.columns] if all(col in dataframe.columns for col in self.columns) else dataframe

        work = dataframe.copy()
        posted_series = work["게시일시"] if "게시일시" in work.columns else pd.Series([""] * len(work), index=work.index)
        date_series = work["일자"] if "일자" in work.columns else pd.Series([""] * len(work), index=work.index)
        posted_dt = pd.to_datetime(posted_series, errors="coerce", utc=True)
        fallback_date = pd.to_datetime(date_series, errors="coerce", utc=True)
        work["_sort_dt"] = posted_dt.fillna(fallback_date)
        work = work.sort_values(by="_sort_dt", ascending=False, kind="stable")
        work = work.drop(columns=["_sort_dt"])
        return work[self.columns]

    def _sanitize_for_excel(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        work = dataframe.copy()
        for column in work.columns:
            if column == "URL":
                continue
            if work[column].dtype != object:
                continue
            work[column] = work[column].map(self._escape_excel_formula)
        return work

    @staticmethod
    def _escape_excel_formula(value: object) -> object:
        if not isinstance(value, str):
            return value
        if not value or value.startswith("'"):
            return value

        stripped = value.lstrip()
        if stripped and stripped[0] in EXCEL_FORMULA_PREFIXES:
            return f"'{value}"
        return value

    @staticmethod
    def _normalize_url(value: object) -> str:
        return str(value or "").strip()

    def _normalized_url_set(self, series: pd.Series) -> set[str]:
        return {self._normalize_url(v) for v in series if self._normalize_url(v)}
